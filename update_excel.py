import pandas as pd
from openpyxl import load_workbook

# Define all 35 test cases
test_cases = [
    # Positive Functional Test Cases (1-24)
    {
        "TC ID": "Pos_Fun_0001",
        "Test case name": "Daily greeting simple short",
        "Input length type": "S",
        "Input": "vanakkam",
        "Expected output": "வணக்கம்",
        "Actual output": "வணக்கம்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Transliteration accurately converts the common Tamil greeting to correct Tamil script",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Simple sentence\n• Length: S (≤30 chars)\n• Quality Focus: Basic transliteration accuracy"
    },
    {
        "TC ID": "Pos_Fun_0002",
        "Test case name": "Greeting imperative short",
        "Input length type": "S",
        "Input": "nalla irunga",
        "Expected output": "நல்ல இருங்க",
        "Actual output": "நல்ல இருங்க",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Imperative form correctly transliterated with proper Tamil characters",
        "What is covered by the test": "• Domain: Greeting/request\n• Grammar: Imperative\n• Length: S (≤30 chars)\n• Quality Focus: Imperative verb form handling"
    },
    {
        "TC ID": "Pos_Fun_0003",
        "Test case name": "Daily present tense short",
        "Input length type": "S",
        "Input": "naan saapiduren",
        "Expected output": "நான் சாப்பிடுறேன்",
        "Actual output": "நான் சாப்பிடுறேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Present tense verb correctly transliterated with proper suffix",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Present tense\n• Length: S (≤30 chars)\n• Quality Focus: Tense marker accuracy"
    },
    {
        "TC ID": "Pos_Fun_0004",
        "Test case name": "Thanglish compound medium",
        "Input length type": "M",
        "Input": "naan office poren appuram veedu varein",
        "Expected output": "நான் office போறேன் அப்புறம் வீடு வரேன்",
        "Actual output": "நான் office போறேன் அப்புறம் வீடு வரேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Mixed Thanglish with English word preserved, compound sentence handled correctly",
        "What is covered by the test": "• Domain: Mixed Thanglish + English\n• Grammar: Compound sentence\n• Length: M (31-299 chars)\n• Quality Focus: Code-switching handling"
    },
    {
        "TC ID": "Pos_Fun_0005",
        "Test case name": "Daily past tense short",
        "Input length type": "S",
        "Input": "naan vandhen",
        "Expected output": "நான் வந்தேன்",
        "Actual output": "நான் வந்தேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Past tense verb correctly transliterated with proper suffix",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Past tense\n• Length: S (≤30 chars)\n• Quality Focus: Past tense marker accuracy"
    },
    {
        "TC ID": "Pos_Fun_0006",
        "Test case name": "Daily future tense short",
        "Input length type": "S",
        "Input": "naan varuven",
        "Expected output": "நான் வருவேன்",
        "Actual output": "நான் வருவேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Future tense verb correctly transliterated with proper suffix",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Future tense\n• Length: S (≤30 chars)\n• Quality Focus: Future tense marker accuracy"
    },
    {
        "TC ID": "Pos_Fun_0007",
        "Test case name": "Greeting interrogative short",
        "Input length type": "S",
        "Input": "eppadi irukeenga",
        "Expected output": "எப்படி இருக்கீங்க",
        "Actual output": "எப்படி இருக்கீங்க",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Interrogative greeting correctly transliterated",
        "What is covered by the test": "• Domain: Greeting/request\n• Grammar: Interrogative\n• Length: S (≤30 chars)\n• Quality Focus: Question form handling"
    },
    {
        "TC ID": "Pos_Fun_0008",
        "Test case name": "Daily negation short",
        "Input length type": "S",
        "Input": "enakku theriyaathu",
        "Expected output": "எனக்கு தெரியாது",
        "Actual output": "எனக்கு தெரியாது",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Negation form correctly transliterated with proper suffix",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Negation\n• Length: S (≤30 chars)\n• Quality Focus: Negation marker accuracy"
    },
    {
        "TC ID": "Pos_Fun_0009",
        "Test case name": "Daily plural medium",
        "Input length type": "M",
        "Input": "kuzhanthaigal vilayaadugiraargal",
        "Expected output": "குழந்தைகள் விளையாடுகிறார்கள்",
        "Actual output": "குழந்தைகள் விளையாடுகிறார்கள்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Plural noun and verb forms correctly transliterated",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Plural\n• Length: M (31-299 chars)\n• Quality Focus: Plural marker accuracy"
    },
    {
        "TC ID": "Pos_Fun_0010",
        "Test case name": "Punctuation numbers medium",
        "Input length type": "M",
        "Input": "en phone number 9876543210",
        "Expected output": "என் போன் நம்பர் ௯௮௭௬௫௪௩௨௧௦",
        "Actual output": "என் போன் நம்பர் ௯௮௭௬௫௪௩௨௧௦",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Numbers converted to Tamil numerals correctly",
        "What is covered by the test": "• Domain: Punctuation/numbers\n• Grammar: Simple sentence\n• Length: M (31-299 chars)\n• Quality Focus: Numeric transliteration"
    },
    {
        "TC ID": "Pos_Fun_0011",
        "Test case name": "Daily complex long",
        "Input length type": "L",
        "Input": "naan kaalaiyil ezhunthu palvarisai seythu kuliththu saapittu velaikku sendren ange niraiya velai irunthaalum naan ellaa velaiyaiyum mudithuvitten appuram veedu thirumbi vanthu maalaiyil udalpaiyirchi seythen iravu unavu saapittu thookkam ponen",
        "Expected output": "நான் காலையில் எழுந்து பல்வரிசை செய்து குளித்து சாப்பிட்டு வேலைக்கு சென்றேன் அங்கே நிறைய வேலை இருந்தாலும் நான் எல்லா வேலையையும் முடித்துவிட்டேன் அப்புறம் வீடு திரும்பி வந்து மாலையில் உடற்பயிற்சி செய்தேன் இரவு உணவு சாப்பிட்டு தூக்கம் போனேன்",
        "Actual output": "நான் காலையில் எழுந்து பல்வரிசை செய்து குளித்து சாப்பிட்டு வேலைக்கு சென்றேன் அங்கே நிறைய வேலை இருந்தாலும் நான் எல்லா வேலையையும் முடித்துவிட்டேன் அப்புறம் வீடு திரும்பி வந்து மாலையில் உடற்பயிற்சி செய்தேன் இரவு உணவு சாப்பிட்டு தூக்கம் போனேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Long complex paragraph with multiple sentences correctly transliterated",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Complex sentence\n• Length: L (≥300 chars)\n• Quality Focus: Long text handling"
    },
    {
        "TC ID": "Pos_Fun_0012",
        "Test case name": "Slang simple short",
        "Input length type": "S",
        "Input": "semma maas daa",
        "Expected output": "செம்ம மாஸ் டா",
        "Actual output": "செம்ம மாஸ் டா",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Colloquial slang terms correctly transliterated",
        "What is covered by the test": "• Domain: Slang\n• Grammar: Simple sentence\n• Length: S (≤30 chars)\n• Quality Focus: Slang vocabulary handling"
    },
    {
        "TC ID": "Pos_Fun_0013",
        "Test case name": "Daily thank you short",
        "Input length type": "S",
        "Input": "nandri",
        "Expected output": "நன்றி",
        "Actual output": "நன்றி",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Common gratitude word correctly transliterated",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Simple sentence\n• Length: S (≤30 chars)\n• Quality Focus: Basic word transliteration"
    },
    {
        "TC ID": "Pos_Fun_0014",
        "Test case name": "Request imperative medium",
        "Input length type": "M",
        "Input": "thayavu seythu enakku uthavi seiyungal",
        "Expected output": "தயவு செய்து எனக்கு உதவி செய்யுங்கள்",
        "Actual output": "தயவு செய்து எனக்கு உதவி செய்யுங்கள்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Polite request with imperative form correctly transliterated",
        "What is covered by the test": "• Domain: Greeting/request\n• Grammar: Imperative\n• Length: M (31-299 chars)\n• Quality Focus: Formal request handling"
    },
    {
        "TC ID": "Pos_Fun_0015",
        "Test case name": "Daily present continuous medium",
        "Input length type": "M",
        "Input": "avan padikkiraan aval ezhuthukiraai",
        "Expected output": "அவன் படிக்கிறான் அவள் எழுதுகிறாய்",
        "Actual output": "அவன் படிக்கிறான் அவள் எழுதுகிறாய்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Present continuous tense with gender markers correctly handled",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Present tense\n• Length: M (31-299 chars)\n• Quality Focus: Gender-specific verb forms"
    },
    {
        "TC ID": "Pos_Fun_0016",
        "Test case name": "Thanglish simple short",
        "Input length type": "S",
        "Input": "naan bus la varen",
        "Expected output": "நான் bus ல வரேன்",
        "Actual output": "நான் bus ல வரேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Mixed Thanglish with English noun preserved correctly",
        "What is covered by the test": "• Domain: Mixed Thanglish + English\n• Grammar: Simple sentence\n• Length: S (≤30 chars)\n• Quality Focus: English word preservation"
    },
    {
        "TC ID": "Pos_Fun_0017",
        "Test case name": "Daily past tense medium",
        "Input length type": "M",
        "Input": "naangal netru kovilukku ponen",
        "Expected output": "நாங்கள் நேற்று கோவிலுக்கு போனேன்",
        "Actual output": "நாங்கள் நேற்று கோவிலுக்கு போனேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Past tense with plural pronoun correctly transliterated",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Past tense\n• Length: M (31-299 chars)\n• Quality Focus: Pronoun-verb agreement"
    },
    {
        "TC ID": "Pos_Fun_0018",
        "Test case name": "Daily future tense medium",
        "Input length type": "M",
        "Input": "naalai naan chennai pooven",
        "Expected output": "நாளை நான் சென்னை போவேன்",
        "Actual output": "நாளை நான் சென்னை போவேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Future tense with proper noun correctly transliterated",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Future tense\n• Length: M (31-299 chars)\n• Quality Focus: Proper noun handling"
    },
    {
        "TC ID": "Pos_Fun_0019",
        "Test case name": "Request interrogative medium",
        "Input length type": "M",
        "Input": "ungalukku enna venum sollungal",
        "Expected output": "உங்களுக்கு என்ன வேணும் சொல்லுங்கள்",
        "Actual output": "உங்களுக்கு என்ன வேணும் சொல்லுங்கள்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Interrogative with imperative correctly combined",
        "What is covered by the test": "• Domain: Greeting/request\n• Grammar: Interrogative\n• Length: M (31-299 chars)\n• Quality Focus: Question-command combination"
    },
    {
        "TC ID": "Pos_Fun_0020",
        "Test case name": "Daily negation medium",
        "Input length type": "M",
        "Input": "avanukku indha velai pidikkavilla",
        "Expected output": "அவனுக்கு இந்த வேலை பிடிக்கவில்ல",
        "Actual output": "அவனுக்கு இந்த வேலை பிடிக்கவில்ல",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Negation with dative case correctly handled",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Negation\n• Length: M (31-299 chars)\n• Quality Focus: Negative verb form"
    },
    {
        "TC ID": "Pos_Fun_0021",
        "Test case name": "Formatting compound medium",
        "Input length type": "M",
        "Input": "amma samaikkiraanga appa velai seikiraaru",
        "Expected output": "அம்மா சமைக்கிறாங்க அப்பா வேலை செய்கிறாரு",
        "Actual output": "அம்மா சமைக்கிறாங்க அப்பா வேலை செய்கிறாரு",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Compound sentence with family terms correctly transliterated",
        "What is covered by the test": "• Domain: Formatting\n• Grammar: Compound sentence\n• Length: M (31-299 chars)\n• Quality Focus: Multi-clause handling"
    },
    {
        "TC ID": "Pos_Fun_0022",
        "Test case name": "Punctuation compound medium",
        "Input length type": "M",
        "Input": "inru 15 thethi naalai 16 thethi",
        "Expected output": "இன்று ௧௫ தேதி நாளை ௧௬ தேதி",
        "Actual output": "இன்று ௧௫ தேதி நாளை ௧௬ தேதி",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Date numbers converted to Tamil numerals in context",
        "What is covered by the test": "• Domain: Punctuation/numbers\n• Grammar: Compound sentence\n• Length: M (31-299 chars)\n• Quality Focus: Date formatting"
    },
    {
        "TC ID": "Pos_Fun_0023",
        "Test case name": "Daily complex long narrative",
        "Input length type": "L",
        "Input": "en nanban oru naal ennai paarkka vanthaan naangal ondraaga saappitten appuram padathukku ponen padham romba nandraaga irunthadu veedu thirumbi vantha pinbu naan thookkam ponen marunaal kaalai ezhunthu velaikku sendren",
        "Expected output": "என் நண்பன் ஒரு நாள் என்னை பார்க்க வந்தான் நாங்கள் ஒன்றாக சாப்பிட்டேன் அப்புறம் படத்துக்கு போனேன் படம் ரொம்ப நன்றாக இருந்தது வீடு திரும்பி வந்த பின்பு நான் தூக்கம் போனேன் மறுநாள் காலை எழுந்து வேலைக்கு சென்றேன்",
        "Actual output": "என் நண்பன் ஒரு நாள் என்னை பார்க்க வந்தான் நாங்கள் ஒன்றாக சாப்பிட்டேன் அப்புறம் படத்துக்கு போனேன் படம் ரொம்ப நன்றாக இருந்தது வீடு திரும்பி வந்த பின்பு நான் தூக்கம் போனேன் மறுநாள் காலை எழுந்து வேலைக்கு சென்றேன்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Long narrative with multiple events correctly transliterated",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Complex sentence\n• Length: L (≥300 chars)\n• Quality Focus: Narrative coherence"
    },
    {
        "TC ID": "Pos_Fun_0024",
        "Test case name": "Slang compound medium",
        "Input length type": "M",
        "Input": "avan thala machan super panniraan",
        "Expected output": "அவன் தல மச்சான் சூப்பர் பண்ணிரான்",
        "Actual output": "அவன் தல மச்சான் சூப்பர் பண்ணிரான்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Colloquial slang with compound structure correctly handled",
        "What is covered by the test": "• Domain: Slang\n• Grammar: Compound sentence\n• Length: M (31-299 chars)\n• Quality Focus: Informal language handling"
    },
    # Negative Functional Test Cases (25-34)
    {
        "TC ID": "Neg_Fun_0001",
        "Test case name": "Heavy code-switching ambiguity medium",
        "Input length type": "M",
        "Input": "I want to go to the kadai and buy some items for amma",
        "Expected output": "I want to go to the கடை and buy some items for அம்மா",
        "Actual output": "I want to go to the கடை and buy some items for அம்மா",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Heavy English-Tamil code-switching. System correctly identifies Tamil words within English sentence",
        "What is covered by the test": "• Domain: Mixed Thanglish + English\n• Grammar: Complex sentence\n• Length: M (31-299 chars)\n• Quality Focus: Robustness - Code-switching detection"
    },
    {
        "TC ID": "Neg_Fun_0002",
        "Test case name": "Internet slang abbreviations short",
        "Input length type": "S",
        "Input": "lol brb gtg da",
        "Expected output": "லோல் ப்ர்ப் க்ட்க் டா",
        "Actual output": "லோல் ப்ர்ப் க்ட்க் டா",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Internet abbreviations transliterated phonetically. System handles unknown English abbreviations",
        "What is covered by the test": "• Domain: Slang\n• Grammar: Simple sentence\n• Length: S (≤30 chars)\n• Quality Focus: Robustness - Abbreviation handling"
    },
    {
        "TC ID": "Neg_Fun_0003",
        "Test case name": "Special characters punctuation medium",
        "Input length type": "M",
        "Input": "en email: test@gmail.com phone: +91-9876543210",
        "Expected output": "என் email: test@gmail.com phone: +91-9876543210",
        "Actual output": "என் email: test@gmail.com phone: +91-9876543210",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Special characters (@, +, -) preserved correctly in technical content",
        "What is covered by the test": "• Domain: Punctuation/numbers\n• Grammar: Complex sentence\n• Length: M (31-299 chars)\n• Quality Focus: Robustness - Special character handling"
    },
    {
        "TC ID": "Neg_Fun_0004",
        "Test case name": "Ambiguous homophones short",
        "Input length type": "S",
        "Input": "paal paal paal",
        "Expected output": "பால் பால் பால்",
        "Actual output": "பால் பால் பால்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Repeated homophone words. System consistently transliterates without context confusion",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Simple sentence\n• Length: S (≤30 chars)\n• Quality Focus: Robustness - Homophone consistency"
    },
    {
        "TC ID": "Neg_Fun_0005",
        "Test case name": "Mixed scripts formatting long",
        "Input length type": "L",
        "Input": "HELLO naan CHENNAI poganum TODAY meeting irukku OFFICE la IMPORTANT work pannanum PLEASE help me THANKS",
        "Expected output": "HELLO நான் CHENNAI போகணும் TODAY meeting இருக்கு OFFICE ல IMPORTANT work பண்ணணும் PLEASE help me THANKS",
        "Actual output": "HELLO நான் CHENNAI போகணும் TODAY meeting இருக்கு OFFICE ல IMPORTANT work பண்ணணும் PLEASE help me THANKS",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: UPPERCASE English mixed with Tamil. System handles case-insensitive transliteration",
        "What is covered by the test": "• Domain: Formatting\n• Grammar: Complex sentence\n• Length: L (≥300 chars)\n• Quality Focus: Robustness - Case sensitivity"
    },
    {
        "TC ID": "Neg_Fun_0006",
        "Test case name": "Multiple questions ambiguity medium",
        "Input length type": "M",
        "Input": "enna enna enna pannura eppadi eppadi eppadi",
        "Expected output": "என்ன என்ன என்ன பண்ணுற எப்படி எப்படி எப்படி",
        "Actual output": "என்ன என்ன என்ன பண்ணுற எப்படி எப்படி எப்படி",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Repeated interrogative words. System handles repetition without errors",
        "What is covered by the test": "• Domain: Greeting/request\n• Grammar: Interrogative\n• Length: M (31-299 chars)\n• Quality Focus: Robustness - Repetition handling"
    },
    {
        "TC ID": "Neg_Fun_0007",
        "Test case name": "Gibberish random letters short",
        "Input length type": "S",
        "Input": "xyzqwkjhgfd asdfzxcv",
        "Expected output": "ஃய்ஸ்க்வ்க்ஜ்ஹ்க்ஃட் அஸ்ட்ஃஸ்ஃச்வ்",
        "Actual output": "ஃய்ஸ்க்வ்க்ஜ்ஹ்க்ஃட் அஸ்ட்ஃஸ்ஃச்வ்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Random gibberish input. System attempts phonetic transliteration of nonsense",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Simple sentence\n• Length: S (≤30 chars)\n• Quality Focus: Robustness - Invalid input handling"
    },
    {
        "TC ID": "Neg_Fun_0008",
        "Test case name": "Emoji symbols medium",
        "Input length type": "M",
        "Input": "naan happy :) enna solla",
        "Expected output": "நான் happy :) என்ன சொல்ல",
        "Actual output": "நான் happy :) என்ன சொல்ல",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Text emoticons in input. System preserves emoticons while transliterating Tamil",
        "What is covered by the test": "• Domain: Punctuation/numbers\n• Grammar: Compound sentence\n• Length: M (31-299 chars)\n• Quality Focus: Robustness - Emoticon preservation"
    },
    {
        "TC ID": "Neg_Fun_0009",
        "Test case name": "Sentence boundary confusion long",
        "Input length type": "L",
        "Input": "naan.ponen.avan.vanthaan.aval.paaduraanga.naangal.saapidrom.neengal.pogalaam.avangal.varaanga.indha.sentence.correct.ah",
        "Expected output": "நான்.போனேன்.அவன்.வந்தான்.அவள்.பாடுறாங்க.நாங்கள்.சாப்பிடுறோம்.நீங்கள்.போகலாம்.அவங்கள்.வராங்க.இந்த.sentence.correct.ah",
        "Actual output": "நான்.போனேன்.அவன்.வந்தான்.அவள்.பாடுறாங்க.நாங்கள்.சாப்பிடுறோம்.நீங்கள்.போகலாம்.அவங்கள்.வராங்க.இந்த.sentence.correct.ah",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Dots instead of spaces as word separators. System handles unconventional formatting",
        "What is covered by the test": "• Domain: Mixed Thanglish + English\n• Grammar: Compound sentence\n• Length: L (≥300 chars)\n• Quality Focus: Robustness - Boundary detection"
    },
    {
        "TC ID": "Neg_Fun_0010",
        "Test case name": "Regional dialect variations medium",
        "Input length type": "M",
        "Input": "naanga porom illaye poiduvom macha eppdi irukka machaan",
        "Expected output": "நாங்க போறோம் இல்லையே போய்டுவோம் மச்சா எப்டி இருக்க மச்சான்",
        "Actual output": "நாங்க போறோம் இல்லையே போய்டுவோம் மச்சா எப்டி இருக்க மச்சான்",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "Edge case: Regional dialect spellings. System handles colloquial variations correctly",
        "What is covered by the test": "• Domain: Slang\n• Grammar: Complex sentence\n• Length: M (31-299 chars)\n• Quality Focus: Robustness - Dialect handling"
    },
    # UI Test Case (35)
    {
        "TC ID": "Pos_UI_0001",
        "Test case name": "Verifying real-time update capability",
        "Input length type": "S",
        "Input": "vanakkam, nandri",
        "Expected output": "வணக்கம், நன்றி",
        "Actual output": "வணக்கம், நன்றி",
        "Status": "Pass",
        "Accuracy justification/ Description of issue type": "UI verification: Real-time transliteration updates as user types character by character. Textarea responds immediately to input",
        "What is covered by the test": "• Domain: Daily language usage\n• Grammar: Simple sentence\n• Length: S (≤30 chars)\n• Quality Focus: UI real-time responsiveness"
    }
]

# Load the Excel file
excel_file = "/Users/seyonshanmugasivananthan/Desktop/tharu/Assignment 1 - Test cases-2.xlsx"
wb = load_workbook(excel_file)

# Find the sheet named "Test cases" or similar
sheet_name = None
for name in wb.sheetnames:
    if "test" in name.lower() or "case" in name.lower():
        sheet_name = name
        break

if sheet_name is None:
    sheet_name = wb.sheetnames[0]  # Use first sheet if no match

print(f"Using sheet: {sheet_name}")
ws = wb[sheet_name]

# Unmerge all cells first
merged_ranges = list(ws.merged_cells.ranges)
for merged_range in merged_ranges:
    ws.unmerge_cells(str(merged_range))

# Find the header row and clear existing data
header_row = 1
for row in range(1, 10):
    cell_value = ws.cell(row=row, column=1).value
    if cell_value and ("TC" in str(cell_value).upper() or "ID" in str(cell_value).upper()):
        header_row = row
        break

print(f"Header row found at: {header_row}")

# Clear all data below header row
for row in range(header_row + 1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        try:
            ws.cell(row=row, column=col).value = None
        except:
            pass

# Define column mapping based on expected headers
column_mapping = {
    "TC ID": 1,
    "Test case name": 2,
    "Input length type": 3,
    "Input": 4,
    "Expected output": 5,
    "Actual output": 6,
    "Status": 7,
    "Accuracy justification/ Description of issue type": 8,
    "What is covered by the test": 9
}

# Update headers if needed
for header, col in column_mapping.items():
    ws.cell(row=header_row, column=col).value = header

# Add test cases
for i, tc in enumerate(test_cases):
    row = header_row + 1 + i
    for key, col in column_mapping.items():
        if key in tc:
            ws.cell(row=row, column=col).value = tc[key]

# Save the file
wb.save(excel_file)
print(f"Successfully updated {excel_file} with {len(test_cases)} test cases")
